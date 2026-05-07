"""
Extract texto de PDF/TXT/HTML/JSON/XLSX. Port do RAG PHP (api/upload.php) para Python.

Convenção:
- PDF/XLSX: páginas/sheets numeradas 1-indexed
- TXT/HTML/JSON: documento single-page (page_number=1)
- Strip de tags HTML faz unescape básico mas não preserva estrutura
- XLSX: cada sheet vira "página"; rows × cols renderizados como TSV
"""
from __future__ import annotations

import json
import logging
import os
import re
from html import unescape
from typing import Optional

from .models import ExtractedDocument, PageText

logger = logging.getLogger("rag.extractor")

# Mapping extension → mime_type (lookup conservador, sem usar mimetypes do stdlib)
_EXT_TO_MIME = {
    ".pdf":  "application/pdf",
    ".txt":  "text/plain",
    ".html": "text/html",
    ".htm":  "text/html",
    ".json": "application/json",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls":  "application/vnd.ms-excel",
}


def _detect_mime(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    mime = _EXT_TO_MIME.get(ext)
    if not mime:
        raise ValueError(
            f"extensão não suportada: '{ext}'. "
            f"Permitidos: {sorted(_EXT_TO_MIME.keys())}"
        )
    return mime


def extract_text(file_path: str, mime_type: Optional[str] = None) -> ExtractedDocument:
    """Extrai texto de um arquivo. Detecta mime por extensão se não informado."""
    if not os.path.isfile(file_path):
        raise FileNotFoundError(file_path)

    mime = mime_type or _detect_mime(file_path)

    if mime == "application/pdf":
        return _extract_pdf(file_path)
    if mime == "text/plain":
        return _extract_txt(file_path, mime)
    if mime == "text/html":
        return _extract_html(file_path, mime)
    if mime == "application/json":
        return _extract_json(file_path, mime)
    if mime in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return _extract_xlsx(file_path, mime)

    raise ValueError(f"mime_type não suportado: {mime}")


def _extract_pdf(file_path: str) -> ExtractedDocument:
    """PDF via pdfplumber. Cada página vira PageText."""
    import pdfplumber  # lazy: dep pesada

    pages: list[PageText] = []
    metadata: dict = {}
    with pdfplumber.open(file_path) as pdf:
        meta = pdf.metadata or {}
        # pdfplumber retorna dict com chaves capitalizadas tipo 'Author', 'Title'
        metadata = {k.lower(): str(v) for k, v in meta.items() if v}
        for i, page in enumerate(pdf.pages, start=1):
            txt = (page.extract_text() or "").strip()
            if txt:
                pages.append(PageText(page_number=i, content=txt))
    full = "\n\n".join(p.content for p in pages)
    return ExtractedDocument(
        full_text=full,
        pages=pages,
        page_count=len(pages),
        mime_type="application/pdf",
        metadata=metadata,
    )


def _extract_txt(file_path: str, mime: str) -> ExtractedDocument:
    """Texto plano. Single-page."""
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        content = f.read().strip()
    pages = [PageText(page_number=1, content=content)] if content else []
    return ExtractedDocument(
        full_text=content,
        pages=pages,
        page_count=1 if content else 0,
        mime_type=mime,
    )


def _extract_html(file_path: str, mime: str) -> ExtractedDocument:
    """HTML: strip tags + unescape entities. Single-page."""
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        raw = f.read()
    # Remove script/style + content
    raw = re.sub(r"<(script|style)[^>]*>.*?</\1>", "", raw, flags=re.S | re.I)
    # Strip tags
    text = re.sub(r"<[^>]+>", " ", raw)
    # Normaliza whitespace + unescape
    text = re.sub(r"\s+", " ", unescape(text)).strip()
    pages = [PageText(page_number=1, content=text)] if text else []
    return ExtractedDocument(
        full_text=text,
        pages=pages,
        page_count=1 if text else 0,
        mime_type=mime,
    )


def _extract_json(file_path: str, mime: str) -> ExtractedDocument:
    """JSON pretty-printed (text). Single-page."""
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        try:
            data = json.load(f)
            text = json.dumps(data, ensure_ascii=False, indent=2)
        except json.JSONDecodeError as e:
            logger.warning("rag.extractor: JSON inválido em %s — caindo para raw read: %s", file_path, e)
            f.seek(0)
            text = f.read().strip()
    pages = [PageText(page_number=1, content=text)] if text else []
    return ExtractedDocument(
        full_text=text,
        pages=pages,
        page_count=1 if text else 0,
        mime_type=mime,
    )


def _extract_xlsx(file_path: str, mime: str) -> ExtractedDocument:
    """XLSX/XLS via openpyxl. Cada sheet vira "página" 1-indexed."""
    from openpyxl import load_workbook  # lazy

    wb = load_workbook(file_path, read_only=True, data_only=True)
    pages: list[PageText] = []
    metadata: dict = {"sheet_names": list(wb.sheetnames)}
    try:
        for i, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            rows: list[str] = [f"# Sheet: {sheet_name}"]
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                if any(cells):
                    rows.append("\t".join(cells))
            sheet_text = "\n".join(rows).strip()
            if sheet_text:
                pages.append(PageText(page_number=i, content=sheet_text))
    finally:
        # Em modo read_only, openpyxl mantém o arquivo aberto até wb.close().
        # Sem isso, no Windows o tempfile não pode ser deletado pelo caller.
        wb.close()
    full = "\n\n".join(p.content for p in pages)
    return ExtractedDocument(
        full_text=full,
        pages=pages,
        page_count=len(pages),
        mime_type=mime,
        metadata=metadata,
    )
